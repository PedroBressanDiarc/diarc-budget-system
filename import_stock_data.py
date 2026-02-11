#!/usr/bin/env python3
"""
Script para importar dados da planilha de estoque para o banco de dados
"""
import openpyxl
import json
import sys

def parse_excel_to_json(excel_file):
    """Converte a planilha Excel em JSON para importação"""
    
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    all_items = []
    
    # Mapear categorias por aba
    category_map = {
        'ESTOQUE': 'Estoque Geral',
        'ESTOQUE - ABRASIVOS': 'Abrasivos',
        'ESTOQUE - EPI': 'EPI',
        'ESTOQUE METALURGICA': 'Metalúrgica',
        'ESTOQUE - IMPORTAÇÃO': 'Importação',
        'PEÇAS - BONPLAND': 'Peças - Bonpland',
        'PEÇAS - LANG': 'Peças - Lang',
    }
    
    for sheet_name in wb.sheetnames:
        if sheet_name == 'TOTAL ESTOQUE':
            continue  # Pular aba de totais
            
        print(f"Processando aba: {sheet_name}")
        ws = wb[sheet_name]
        category = category_map.get(sheet_name, sheet_name)
        
        # Ler cabeçalhos
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(1, col).value
            headers.append(str(cell_value) if cell_value else f"Col{col}")
        
        # Processar linhas de dados
        for row in range(2, ws.max_row + 1):
            row_data = {}
            has_data = False
            
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row, col).value
                if cell_value is not None and str(cell_value).strip():
                    has_data = True
                row_data[headers[col-1]] = cell_value
            
            if not has_data:
                continue
            
            # Extrair informações do item
            item_name = None
            quantity = 0
            unit_price = 0
            unit = "un"
            
            # Identificar nome do item (primeira coluna com texto)
            for key, value in row_data.items():
                if value and isinstance(value, str) and len(value) > 3:
                    item_name = value
                    break
            
            if not item_name or item_name.startswith('ESTOQUE'):
                continue  # Pular cabeçalhos e linhas vazias
            
            # Identificar quantidade e preço baseado na estrutura da planilha
            if 'QTDE' in str(row_data.get('Col2', '')).upper():
                continue  # Pular linha de cabeçalho
            
            # Tentar extrair quantidade
            for key in ['Col2', 'UN', 'Quantidade']:
                val = row_data.get(key)
                if val and isinstance(val, (int, float)):
                    quantity = float(val)
                    break
            
            # Tentar extrair preço unitário
            for key in ['Col3', 'Col4', 'Valor unitario', 'Preço']:
                val = row_data.get(key)
                if val and isinstance(val, (int, float)) and val > 0:
                    # Se for muito grande, pode ser valor total, não unitário
                    if val < 10000:
                        unit_price = float(val)
                        break
            
            # Determinar unidade baseada na categoria
            if category == 'Metalúrgica':
                unit = 'kg'
            elif 'PEÇAS' in category.upper():
                unit = 'm3'
            else:
                unit = 'un'
            
            # Criar item
            item = {
                'name': item_name,
                'category': category,
                'quantity': quantity,
                'unitPrice': unit_price,
                'defaultUnit': unit,
                'notes': f'Importado da planilha - Aba: {sheet_name}'
            }
            
            all_items.append(item)
    
    return all_items

def generate_sql_inserts(items):
    """Gera comandos SQL INSERT para os itens"""
    
    sql_commands = []
    sql_commands.append("-- Importação de dados da planilha de estoque")
    sql_commands.append("-- Execute este script no banco de dados MySQL/TiDB")
    sql_commands.append("")
    
    for item in items:
        name = item['name'].replace("'", "''")  # Escapar aspas simples
        category = item['category'].replace("'", "''")
        notes = item.get('notes', '').replace("'", "''")
        
        sql = f"""INSERT INTO items (name, category, quantity, unitPrice, defaultUnit, notes, active, createdBy, createdAt, updatedAt) 
VALUES ('{name}', '{category}', {item['quantity']}, {item['unitPrice']}, '{item['defaultUnit']}', '{notes}', 1, 1, NOW(), NOW());"""
        
        sql_commands.append(sql)
    
    return "\n".join(sql_commands)

def main():
    if len(sys.argv) < 2:
        print("Uso: python3 import_stock_data.py <arquivo_excel>")
        print("Exemplo: python3 import_stock_data.py /home/ubuntu/upload/Cópiade12ESTOQUE-2025-DEZEMBRO.xlsx")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    
    print(f"Lendo planilha: {excel_file}")
    items = parse_excel_to_json(excel_file)
    
    print(f"\nTotal de itens encontrados: {len(items)}")
    
    # Salvar JSON
    json_file = '/home/ubuntu/diarc-budget-system-new/stock_data.json'
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(items, f, ensure_ascii=False, indent=2)
    print(f"Dados salvos em JSON: {json_file}")
    
    # Gerar SQL
    sql_file = '/home/ubuntu/diarc-budget-system-new/import_stock.sql'
    sql_content = generate_sql_inserts(items)
    with open(sql_file, 'w', encoding='utf-8') as f:
        f.write(sql_content)
    print(f"Script SQL gerado: {sql_file}")
    
    # Mostrar primeiros 5 itens como exemplo
    print("\nPrimeiros 5 itens:")
    for i, item in enumerate(items[:5], 1):
        print(f"{i}. {item['name']} - {item['quantity']} {item['defaultUnit']} - R$ {item['unitPrice']}")
    
    print("\n✅ Importação concluída!")
    print(f"Execute o arquivo {sql_file} no banco de dados para importar os itens.")

if __name__ == '__main__':
    main()
